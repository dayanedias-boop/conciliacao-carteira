"""
Conciliação de Carteira — Direitos Creditórios
App Streamlit — interface web para rodar sem instalar nada
"""

import io
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ── Configuração da página ─────────────────────────────────────────────────
st.set_page_config(
    page_title="Conciliação — Carteira DC",
    page_icon="🔄",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── CSS customizado ────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

.main { background: #F7F9FC; }

.hero {
    background: linear-gradient(135deg, #1F3864 0%, #2F5496 100%);
    border-radius: 16px;
    padding: 40px 48px;
    color: white;
    margin-bottom: 32px;
}
.hero h1 { font-size: 28px; font-weight: 700; margin: 0 0 8px 0; letter-spacing: -0.5px; }
.hero p  { font-size: 15px; opacity: 0.8; margin: 0; }

.card {
    background: white;
    border-radius: 12px;
    padding: 24px 28px;
    border: 1px solid #E8EDF5;
    margin-bottom: 16px;
    box-shadow: 0 1px 4px rgba(31,56,100,0.06);
}
.card h3 { font-size: 14px; font-weight: 600; color: #1F3864; margin: 0 0 4px 0; text-transform: uppercase; letter-spacing: 0.5px; }
.card p  { font-size: 13px; color: #6B7A99; margin: 0; }

.stat-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin: 24px 0; }
.stat {
    border-radius: 12px;
    padding: 20px 22px;
    border: 1px solid #E8EDF5;
}
.stat .num  { font-size: 28px; font-weight: 700; line-height: 1; }
.stat .lbl  { font-size: 12px; font-weight: 500; margin-top: 6px; opacity: 0.75; }
.stat.verde  { background: #F0FAF4; border-color: #C3E6CE; }
.stat.verde .num { color: #1A7F3C; }
.stat.verde .lbl { color: #1A7F3C; }
.stat.verm   { background: #FFF5F5; border-color: #FFC9C9; }
.stat.verm .num  { color: #C92A2A; }
.stat.verm .lbl  { color: #C92A2A; }
.stat.laran  { background: #FFF8F0; border-color: #FFCC80; }
.stat.laran .num { color: #E65100; }
.stat.laran .lbl { color: #E65100; }
.stat.cinza  { background: #F7F9FC; border-color: #D0D9E8; }
.stat.cinza .num { color: #364F6B; }
.stat.cinza .lbl { color: #364F6B; }

.step {
    display: flex; align-items: flex-start; gap: 16px;
    padding: 16px 0; border-bottom: 1px solid #F0F4FA;
}
.step:last-child { border-bottom: none; }
.step-num {
    width: 32px; height: 32px; border-radius: 50%;
    background: #1F3864; color: white;
    font-size: 14px; font-weight: 700;
    display: flex; align-items: center; justify-content: center;
    flex-shrink: 0; margin-top: 2px;
}
.step-txt h4 { font-size: 14px; font-weight: 600; color: #1F3864; margin: 0 0 4px 0; }
.step-txt p  { font-size: 13px; color: #6B7A99; margin: 0; }

.badge {
    display: inline-block; padding: 3px 10px; border-radius: 20px;
    font-size: 12px; font-weight: 600;
}
.badge-novo  { background: #E8F5E9; color: #1A7F3C; }
.badge-saiu  { background: #FFEBEE; color: #C92A2A; }
.badge-alt   { background: #FFF3E0; color: #E65100; }
.badge-sem   { background: #ECEFF1; color: #455A64; }

.financeiro {
    display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px; margin: 16px 0;
}
.fin-item {
    background: white; border-radius: 10px; padding: 16px 20px;
    border: 1px solid #E8EDF5;
}
.fin-item .fin-lbl { font-size: 12px; color: #8898AA; font-weight: 500; margin-bottom: 6px; }
.fin-item .fin-val { font-size: 20px; font-weight: 700; color: #1F3864; }
.fin-item .fin-val.pos { color: #1A7F3C; }
.fin-item .fin-val.neg { color: #C92A2A; }

div[data-testid="stFileUploader"] > div {
    border: 2px dashed #C5D3E8 !important;
    border-radius: 12px !important;
    background: #F7F9FC !important;
}
div[data-testid="stFileUploader"] > div:hover {
    border-color: #2F5496 !important;
    background: #EEF3FB !important;
}
.stButton > button {
    background: linear-gradient(135deg, #1F3864, #2F5496) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 12px 32px !important;
    font-size: 15px !important;
    font-weight: 600 !important;
    width: 100% !important;
    letter-spacing: 0.3px;
}
.stButton > button:hover { opacity: 0.92 !important; }

.stDownloadButton > button {
    background: linear-gradient(135deg, #1A7F3C, #2E9E56) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 12px 32px !important;
    font-size: 15px !important;
    font-weight: 600 !important;
    width: 100% !important;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# LÓGICA DE CONCILIAÇÃO (igual ao script original)
# ══════════════════════════════════════════════════════════════════════════════
COLUNAS_CHAVE = [
    "id_cobranca", "valor_cobranca", "valor_pago",
    "valor_excedente_pago", "status_pagamento", "contas_a_receber",
]

C_AZE = "1F3864"; C_AZM = "2F5496"; C_BRA = "FFFFFF"
C_CZC = "F2F2F2"; C_VER = "E2EFDA"; C_VRM = "FFDCE1"
C_LAR = "FCE4D6"; C_AMA = "FFF2CC"
FMT_R = "#,##0.00"; FMT_RD = "#,##0.00;[RED]-#,##0.00"; FMT_N = "#,##0"
COR_MOV = {"🟢 NOVO": C_VER, "🔴 SAIU DA CARTEIRA": C_VRM,
           "🟡 ALTERADO": C_LAR, "⚪ SEM MOVIMENTAÇÃO": C_CZC}

def fill(c):    return PatternFill("solid", fgColor=c)
def borda():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def alin(h="center", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
def hdr(cell, val, bg=C_AZM, size=10):
    cell.value = val
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=size)
    cell.fill = fill(bg); cell.alignment = alin(wrap=True); cell.border = borda()
def cel(cell, val=None, bg=C_BRA, fmt=None, align="center"):
    if val is not None: cell.value = val
    cell.font = Font(name="Arial", size=10)
    cell.fill = fill(bg); cell.alignment = alin(align); cell.border = borda()
    if fmt: cell.number_format = fmt

def detectar_aba(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    for nome in wb.sheetnames:
        ws = wb[nome]
        try:
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if "id_cobranca" in headers:
                wb.close()
                return nome
        except StopIteration:
            continue
    wb.close()
    raise ValueError(f"Nenhuma aba com 'id_cobranca' encontrada. Abas disponíveis: {wb.sheetnames}")

def ler_base(file_bytes, label):
    aba = detectar_aba(file_bytes)
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=aba,
                       dtype={"id_cobranca": str}, engine="openpyxl")
    ausentes = [c for c in COLUNAS_CHAVE if c not in df.columns]
    if ausentes:
        raise ValueError(f"Colunas ausentes em '{label}': {ausentes}")
    for col in ["valor_cobranca", "valor_pago", "valor_excedente_pago", "contas_a_receber"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    df["status_pagamento"] = df["status_pagamento"].fillna("").astype(str)
    return df, aba

def conciliar(df_ant, df_atu):
    ant = df_ant[COLUNAS_CHAVE].copy()
    atu = df_atu[COLUNAS_CHAVE].copy()
    merged = pd.merge(ant, atu, on="id_cobranca", how="outer", suffixes=("_ant", "_atu"))
    for sfx in ("_ant", "_atu"):
        for col in ["valor_cobranca", "valor_pago", "valor_excedente_pago", "contas_a_receber"]:
            merged[f"{col}{sfx}"] = merged[f"{col}{sfx}"].fillna(0)
        merged[f"status_pagamento{sfx}"] = merged[f"status_pagamento{sfx}"].fillna("")

    def classificar(row):
        tem_ant = row["valor_cobranca_ant"] != 0 or row["status_pagamento_ant"] != ""
        tem_atu = row["valor_cobranca_atu"] != 0 or row["status_pagamento_atu"] != ""
        if not tem_ant and tem_atu:  return "🟢 NOVO"
        if tem_ant and not tem_atu:  return "🔴 SAIU DA CARTEIRA"
        if any([
            row["status_pagamento_ant"] != row["status_pagamento_atu"],
            abs(row["valor_pago_ant"]          - row["valor_pago_atu"])          > 0.001,
            abs(row["valor_excedente_pago_ant"] - row["valor_excedente_pago_atu"]) > 0.001,
            abs(row["contas_a_receber_ant"]     - row["contas_a_receber_atu"])     > 0.001,
        ]): return "🟡 ALTERADO"
        return "⚪ SEM MOVIMENTAÇÃO"

    merged["tipo_movimentacao"] = merged.apply(classificar, axis=1)
    merged["var_valor_pago"]       = merged["valor_pago_atu"]          - merged["valor_pago_ant"]
    merged["var_excedente_pago"]   = merged["valor_excedente_pago_atu"] - merged["valor_excedente_pago_ant"]
    merged["var_contas_a_receber"] = merged["contas_a_receber_atu"]    - merged["contas_a_receber_ant"]

    def campos_alt(row):
        if row["tipo_movimentacao"] != "🟡 ALTERADO": return "—"
        p = []
        if row["status_pagamento_ant"] != row["status_pagamento_atu"]: p.append("status_pagamento")
        if abs(row["valor_pago_ant"] - row["valor_pago_atu"]) > 0.001: p.append("valor_pago")
        if abs(row["valor_excedente_pago_ant"] - row["valor_excedente_pago_atu"]) > 0.001: p.append("valor_excedente_pago")
        if abs(row["contas_a_receber_ant"] - row["contas_a_receber_atu"]) > 0.001: p.append("contas_a_receber")
        return " | ".join(p)

    merged["campos_alterados"] = merged.apply(campos_alt, axis=1)
    return merged

def gravar_aba_excel(wb, nome, df_d, titulo, tab_color, colunas):
    ws = wb.create_sheet(nome)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A4"
    lc = get_column_letter(len(colunas))
    ws.row_dimensions[1].height = 44
    ws.merge_cells(f"A1:{lc}1")
    c = ws["A1"]
    c.value = titulo
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    c.fill = fill(C_AZE); c.alignment = alin(wrap=True)
    ws.row_dimensions[2].height = 20
    ws.merge_cells(f"A2:{lc}2")
    c = ws["A2"]
    c.value = f"Total de registros: {len(df_d):,}"
    c.font = Font(name="Arial", italic=True, color="595959", size=9)
    c.fill = fill(C_AMA); c.alignment = alin("left", indent=1); c.border = borda()
    ws.row_dimensions[3].height = 42
    for ci, (_, hd, w, _, _) in enumerate(colunas, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        hdr(ws.cell(3, ci), hd)
    for ri, (_, row) in enumerate(df_d.iterrows(), 4):
        ws.row_dimensions[ri].height = 18
        tipo = row.get("tipo_movimentacao", "")
        bg = COR_MOV.get(tipo, C_BRA if ri % 2 == 0 else C_CZC)
        for ci, (col_df, _, _, fmt, align) in enumerate(colunas, 1):
            c2 = ws.cell(ri, ci)
            val = row.get(col_df)
            if pd.isna(val) if not isinstance(val, str) else val == "nan": val = None
            cel(c2, val, bg=bg, fmt=fmt, align=align)
    return ws

def gerar_excel_bytes(conc, df_ant, df_atu, label_ant, label_atu):
    wb = openpyxl.Workbook()
    ws_i = wb.active; ws_i.title = "📋 Instruções"
    ws_i.sheet_view.showGridLines = False
    ws_i.sheet_properties.tabColor = C_AZE
    for col, w in [("A",3),("B",36),("C",54),("D",3)]:
        ws_i.column_dimensions[col].width = w
    ws_i.row_dimensions[1].height = 48
    ws_i.merge_cells("B1:C1")
    c = ws_i["B1"]
    c.value = "CONCILIAÇÃO — CARTEIRA DE DIREITOS CREDITÓRIOS\nGuia de Utilização"
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    c.fill = fill(C_AZE); c.alignment = alin(wrap=True)

    def s(row, txt):
        ws_i.row_dimensions[row].height = 26
        ws_i.merge_cells(f"B{row}:C{row}")
        c = ws_i[f"B{row}"]
        c.value = txt
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = fill(C_AZM); c.alignment = alin("left", indent=1); c.border = borda()

    def it(row, lbl, desc, bg=C_BRA):
        ws_i.row_dimensions[row].height = 22
        cl = ws_i[f"B{row}"]
        cl.value = lbl; cl.font = Font(name="Arial", bold=True, color=C_AZE, size=10)
        cl.fill = fill(bg); cl.alignment = alin("left", indent=1); cl.border = borda()
        cd = ws_i[f"C{row}"]
        cd.value = desc; cd.font = Font(name="Arial", size=10, color="404040")
        cd.fill = fill(bg); cd.alignment = alin("left", indent=1, wrap=True); cd.border = borda()

    r = 3
    s(r, "BASES UTILIZADAS NESTA CONCILIAÇÃO"); r+=1
    it(r, "Mês Anterior", label_ant); r+=1
    it(r, "Mês Atual",    label_atu); r+=1
    r+=1
    s(r, "LEGENDA — TIPOS DE MOVIMENTAÇÃO"); r+=1
    it(r, "🟢 NOVO",             "id_cobranca presente somente no Mês Atual.", C_VER); r+=1
    it(r, "🔴 SAIU DA CARTEIRA", "id_cobranca presente somente no Mês Anterior.", C_VRM); r+=1
    it(r, "🟡 ALTERADO",         "Presente em ambas as bases com diferença em status ou valores.", C_LAR); r+=1
    it(r, "⚪ SEM MOVIMENTAÇÃO", "Presente em ambas as bases sem nenhuma alteração.", C_CZC); r+=1

    # Resumo
    ws_r = wb.create_sheet("📊 Resumo")
    ws_r.sheet_view.showGridLines = False
    ws_r.sheet_properties.tabColor = C_AZE
    for col, w in [("A",3),("B",44),("C",26),("D",3)]:
        ws_r.column_dimensions[col].width = w
    ws_r.row_dimensions[1].height = 48
    ws_r.merge_cells("B1:C1")
    c = ws_r["B1"]
    c.value = f"RESUMO — CONCILIAÇÃO CARTEIRA DE DIREITOS CREDITÓRIOS\n{label_ant}  ×  {label_atu}"
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    c.fill = fill(C_AZE); c.alignment = alin(wrap=True)

    def bloco(rs, titulo, itens):
        ws_r.row_dimensions[rs].height = 28
        ws_r.merge_cells(f"B{rs}:C{rs}")
        c = ws_r[f"B{rs}"]
        c.value = titulo
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = fill(C_AZM); c.alignment = alin(); c.border = borda()
        rr = rs + 1
        for lbl, val, fmt, bg in itens:
            ws_r.row_dimensions[rr].height = 22
            cl = ws_r[f"B{rr}"]
            cl.value = lbl; cl.font = Font(name="Arial", bold=True, color=C_AZE, size=10)
            cl.fill = fill(bg); cl.alignment = alin("left", indent=1); cl.border = borda()
            cv = ws_r[f"C{rr}"]
            cv.value = val; cv.number_format = fmt
            cv.font = Font(name="Arial", bold=True, size=11)
            cv.fill = fill(bg); cv.alignment = alin("right"); cv.border = borda()
            rr += 1
        return rr + 1

    m_nov = conc["tipo_movimentacao"] == "🟢 NOVO"
    m_sai = conc["tipo_movimentacao"] == "🔴 SAIU DA CARTEIRA"
    m_alt = conc["tipo_movimentacao"] == "🟡 ALTERADO"
    def soma(mask, col): return conc.loc[mask, col].fillna(0).sum()

    rr = 3
    rr = bloco(rr, "POSIÇÃO GERAL DA CARTEIRA", [
        (f"Cobranças — {label_ant}",             len(df_ant),                                   FMT_N, C_CZC),
        (f"Cobranças — {label_atu}",             len(df_atu),                                   FMT_N, C_CZC),
        (f"contas_a_receber — {label_ant} (R$)", df_ant["contas_a_receber"].sum(),              FMT_R, C_CZC),
        (f"contas_a_receber — {label_atu} (R$)", df_atu["contas_a_receber"].sum(),              FMT_R, C_CZC),
        ("Variação contas_a_receber (R$)",        df_atu["contas_a_receber"].sum()-df_ant["contas_a_receber"].sum(), FMT_RD, C_AMA),
        (f"valor_pago total — {label_ant} (R$)", df_ant["valor_pago"].sum(),                    FMT_R, C_CZC),
        (f"valor_pago total — {label_atu} (R$)", df_atu["valor_pago"].sum(),                    FMT_R, C_CZC),
    ])
    rr = bloco(rr, "MOVIMENTAÇÕES — QUANTIDADE", [
        ("🟢 NOVO",               m_nov.sum(), FMT_N, C_VER),
        ("🔴 SAIU DA CARTEIRA",   m_sai.sum(), FMT_N, C_VRM),
        ("🟡 ALTERADO",           m_alt.sum(), FMT_N, C_LAR),
        ("⚪ SEM MOVIMENTAÇÃO",   (conc["tipo_movimentacao"]=="⚪ SEM MOVIMENTAÇÃO").sum(), FMT_N, C_CZC),
        ("TOTAL CONCILIADOS",     len(conc),   FMT_N, C_AMA),
    ])
    rr = bloco(rr, "MOVIMENTAÇÕES — contas_a_receber (R$)", [
        ("🟢 CAR Novos (R$)",              soma(m_nov,"contas_a_receber_atu"),  FMT_R,  C_VER),
        ("🔴 CAR Saiu (R$)",               soma(m_sai,"contas_a_receber_ant"),  FMT_R,  C_VRM),
        ("🟡 CAR Alterados Ant. (R$)",     soma(m_alt,"contas_a_receber_ant"),  FMT_R,  C_LAR),
        ("🟡 CAR Alterados Atu. (R$)",     soma(m_alt,"contas_a_receber_atu"),  FMT_R,  C_LAR),
        ("🟡 Variação CAR Alterados (R$)", soma(m_alt,"var_contas_a_receber"),  FMT_RD, C_LAR),
    ])

    # Abas de detalhe
    conc_mov = conc[conc["tipo_movimentacao"] != "⚪ SEM MOVIMENTAÇÃO"].sort_values("tipo_movimentacao")
    COLS_G = [
        ("id_cobranca",             "id_cobrança",                        18, "@",    "center"),
        ("tipo_movimentacao",       "Tipo de\nMovimentação",              26, "@",    "center"),
        ("campos_alterados",        "Campos Alterados",                   34, "@",    "left"),
        ("status_pagamento_ant",    f"status_pag.\n({label_ant})",        22, "@",    "center"),
        ("status_pagamento_atu",    f"status_pag.\n({label_atu})",        22, "@",    "center"),
        ("valor_cobranca_ant",      "valor_cobrança\nAnt. (R$)",          18, FMT_R,  "right"),
        ("valor_cobranca_atu",      "valor_cobrança\nAtu. (R$)",          18, FMT_R,  "right"),
        ("valor_pago_ant",          "valor_pago\nAnt. (R$)",              18, FMT_R,  "right"),
        ("valor_pago_atu",          "valor_pago\nAtu. (R$)",              18, FMT_R,  "right"),
        ("var_valor_pago",          "Var. valor_pago\n(R$)",              18, FMT_RD, "right"),
        ("valor_excedente_pago_ant","excedente_pago\nAnt. (R$)",          20, FMT_R,  "right"),
        ("valor_excedente_pago_atu","excedente_pago\nAtu. (R$)",          20, FMT_R,  "right"),
        ("var_excedente_pago",      "Var. excedente\n(R$)",               18, FMT_RD, "right"),
        ("contas_a_receber_ant",    "contas_a_receber\nAnt. (R$)",        22, FMT_R,  "right"),
        ("contas_a_receber_atu",    "contas_a_receber\nAtu. (R$)",        22, FMT_R,  "right"),
        ("var_contas_a_receber",    "Var. CAR (R$)",                      18, FMT_RD, "right"),
    ]
    gravar_aba_excel(wb,"🔄 Conciliação Geral",conc_mov,
        f"CONCILIAÇÃO GERAL — COBRANÇAS COM MOVIMENTAÇÃO  |  {label_ant} × {label_atu}","833C00",COLS_G)

    conc_alt = conc[conc["tipo_movimentacao"]=="🟡 ALTERADO"].sort_values("campos_alterados")
    COLS_A = [c for c in COLS_G if c[0] not in ("tipo_movimentacao",)]
    gravar_aba_excel(wb,"🟡 Alterados",conc_alt,
        f"ALTERADOS — COBRANÇAS COM MOVIMENTAÇÃO NOS VALORES  |  {label_ant} × {label_atu}","833C00",COLS_A)

    conc_nov = conc[conc["tipo_movimentacao"]=="🟢 NOVO"]
    COLS_N = [
        ("id_cobranca","id_cobrança",18,"@","center"),
        ("valor_cobranca_atu","valor_cobrança (R$)",18,FMT_R,"right"),
        ("status_pagamento_atu","status_pagamento",22,"@","center"),
        ("valor_pago_atu","valor_pago (R$)",18,FMT_R,"right"),
        ("valor_excedente_pago_atu","excedente_pago (R$)",20,FMT_R,"right"),
        ("contas_a_receber_atu","contas_a_receber (R$)",22,FMT_R,"right"),
    ]
    gravar_aba_excel(wb,"🟢 Novos",conc_nov,
        f"NOVOS — Cobranças que entraram em {label_atu}","375623",COLS_N)

    conc_sai = conc[conc["tipo_movimentacao"]=="🔴 SAIU DA CARTEIRA"]
    COLS_S = [
        ("id_cobranca","id_cobrança",18,"@","center"),
        ("valor_cobranca_ant","valor_cobrança (R$)",18,FMT_R,"right"),
        ("status_pagamento_ant","status_pagamento",22,"@","center"),
        ("valor_pago_ant","valor_pago (R$)",18,FMT_R,"right"),
        ("valor_excedente_pago_ant","excedente_pago (R$)",20,FMT_R,"right"),
        ("contas_a_receber_ant","contas_a_receber (R$)",22,FMT_R,"right"),
    ]
    gravar_aba_excel(wb,"🔴 Saíram",conc_sai,
        f"SAÍRAM — Cobranças presentes em {label_ant} e ausentes em {label_atu}","9C0006",COLS_S)

    ordem = ["📋 Instruções","📊 Resumo","🔄 Conciliação Geral","🟡 Alterados","🟢 Novos","🔴 Saíram"]
    wb._sheets.sort(key=lambda s: ordem.index(s.title) if s.title in ordem else 99)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# INTERFACE
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="hero">
    <h1>🔄 Conciliação de Carteira</h1>
    <p>Direitos Creditórios — cruzamento automático entre dois meses</p>
</div>
""", unsafe_allow_html=True)

# Estado da sessão
if "resultado" not in st.session_state:
    st.session_state.resultado = None
if "conc" not in st.session_state:
    st.session_state.conc = None

# ── Upload ─────────────────────────────────────────────────────────────────
col1, col2 = st.columns(2)

with col1:
    st.markdown('<div class="card"><h3>📥 Mês Anterior</h3><p>Base de comparação</p></div>', unsafe_allow_html=True)
    f_ant = st.file_uploader("", type=["xlsx"], key="ant", label_visibility="collapsed")
    if f_ant:
        st.success(f"✓ {f_ant.name}")

with col2:
    st.markdown('<div class="card"><h3>📥 Mês Atual</h3><p>Nova posição da carteira</p></div>', unsafe_allow_html=True)
    f_atu = st.file_uploader("", type=["xlsx"], key="atu", label_visibility="collapsed")
    if f_atu:
        st.success(f"✓ {f_atu.name}")

st.markdown("<br>", unsafe_allow_html=True)

# ── Botão executar ─────────────────────────────────────────────────────────
col_btn, col_gap = st.columns([1, 2])
with col_btn:
    executar = st.button("🔄  Executar Conciliação", disabled=not (f_ant and f_atu))

if executar and f_ant and f_atu:
    with st.spinner("Lendo arquivos e cruzando bases..."):
        try:
            bytes_ant = f_ant.read()
            bytes_atu = f_atu.read()

            df_ant, aba_ant = ler_base(bytes_ant, f_ant.name)
            df_atu, aba_atu = ler_base(bytes_atu, f_atu.name)

            label_ant = Path(f_ant.name).stem
            label_atu = Path(f_atu.name).stem

            conc = conciliar(df_ant, df_atu)

            excel_bytes = gerar_excel_bytes(conc, df_ant, df_atu, label_ant, label_atu)

            st.session_state.resultado  = excel_bytes
            st.session_state.conc       = conc
            st.session_state.df_ant     = df_ant
            st.session_state.df_atu     = df_atu
            st.session_state.label_ant  = label_ant
            st.session_state.label_atu  = label_atu
            st.session_state.nome_saida = f"Conciliacao_{label_atu}.xlsx"

        except Exception as e:
            st.error(f"❌ Erro: {e}")

# ── Resultado ──────────────────────────────────────────────────────────────
if st.session_state.resultado:
    conc      = st.session_state.conc
    df_ant    = st.session_state.df_ant
    df_atu    = st.session_state.df_atu
    label_ant = st.session_state.label_ant
    label_atu = st.session_state.label_atu

    n_nov = (conc["tipo_movimentacao"] == "🟢 NOVO").sum()
    n_sai = (conc["tipo_movimentacao"] == "🔴 SAIU DA CARTEIRA").sum()
    n_alt = (conc["tipo_movimentacao"] == "🟡 ALTERADO").sum()
    n_sem = (conc["tipo_movimentacao"] == "⚪ SEM MOVIMENTAÇÃO").sum()

    st.markdown("---")
    st.markdown("### ✅ Conciliação concluída")

    # Cards de movimentação
    st.markdown(f"""
    <div class="stat-grid">
        <div class="stat verde">
            <div class="num">{n_nov:,}</div>
            <div class="lbl">🟢 Novos</div>
        </div>
        <div class="stat verm">
            <div class="num">{n_sai:,}</div>
            <div class="lbl">🔴 Saíram da Carteira</div>
        </div>
        <div class="stat laran">
            <div class="num">{n_alt:,}</div>
            <div class="lbl">🟡 Alterados</div>
        </div>
        <div class="stat cinza">
            <div class="num">{n_sem:,}</div>
            <div class="lbl">⚪ Sem Movimentação</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Financeiro
    car_ant = df_ant["contas_a_receber"].sum()
    car_atu = df_atu["contas_a_receber"].sum()
    var_car = car_atu - car_ant
    sinal   = "pos" if var_car >= 0 else "neg"

    st.markdown(f"""
    <div class="financeiro">
        <div class="fin-item">
            <div class="fin-lbl">contas_a_receber — {label_ant}</div>
            <div class="fin-val">R$ {car_ant:,.2f}</div>
        </div>
        <div class="fin-item">
            <div class="fin-lbl">contas_a_receber — {label_atu}</div>
            <div class="fin-val">R$ {car_atu:,.2f}</div>
        </div>
        <div class="fin-item">
            <div class="fin-lbl">Variação contas_a_receber</div>
            <div class="fin-val {sinal}">R$ {var_car:+,.2f}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Botão de download
    st.markdown("<br>", unsafe_allow_html=True)
    col_dl, col_g = st.columns([1, 2])
    with col_dl:
        st.download_button(
            label="⬇️  Baixar Excel com resultado",
            data=st.session_state.resultado,
            file_name=st.session_state.nome_saida,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.markdown("---")

    # ── Gráfico: distribuição de status_pagamento_atu ──────────────────────
    st.markdown("### 📊 Distribuição por Status — Mês Atual")

    df_status = df_atu.copy()
    df_status["status_pagamento"] = df_status["status_pagamento"].replace("", "(sem status — em aberto)")
    contagem = df_status["status_pagamento"].value_counts().reset_index()
    contagem.columns = ["Status", "Quantidade"]

    # Cores por status
    CORES_STATUS = {
        "liquidado":                  "#1A7F3C",
        "pagamento_parcial":          "#E65100",
        "original_liquidado":         "#1565C0",
        "(sem status — em aberto)":   "#6B7A99",
    }
    contagem["cor"] = contagem["Status"].map(lambda x: CORES_STATUS.get(x, "#9E9E9E"))

    col_graf, col_tab = st.columns([3, 2])

    with col_graf:
        import plotly.graph_objects as go

        fig = go.Figure(go.Bar(
            x=contagem["Status"],
            y=contagem["Quantidade"],
            marker_color=contagem["cor"].tolist(),
            text=contagem["Quantidade"].apply(lambda v: f"{v:,}"),
            textposition="outside",
            textfont=dict(size=13, family="Inter, sans-serif"),
        ))
        fig.update_layout(
            plot_bgcolor="white",
            paper_bgcolor="white",
            margin=dict(t=20, b=40, l=20, r=20),
            height=360,
            xaxis=dict(
                tickfont=dict(size=12, family="Inter, sans-serif"),
                showgrid=False,
                linecolor="#E8EDF5",
            ),
            yaxis=dict(
                tickfont=dict(size=11, family="Inter, sans-serif"),
                gridcolor="#F0F4FA",
                showgrid=True,
                zeroline=False,
            ),
            bargap=0.35,
        )
        st.plotly_chart(fig, use_container_width=True)

    with col_tab:
        st.markdown("<br>", unsafe_allow_html=True)
        total = contagem["Quantidade"].sum()
        contagem["%"] = (contagem["Quantidade"] / total * 100).round(1).astype(str) + "%"
        st.dataframe(
            contagem[["Status", "Quantidade", "%"]],
            use_container_width=True,
            hide_index=True,
            height=320,
        )

    st.markdown("---")

    # ── Filtros e tabela: Campos Alterados ────────────────────────────────
    st.markdown("### 🔍 Explorar Cobranças Alteradas")

    conc_alt = conc[conc["tipo_movimentacao"] == "🟡 ALTERADO"].copy()

    if conc_alt.empty:
        st.info("Nenhuma cobrança alterada nesta conciliação.")
    else:
        # Opções de campos alterados (multiselect)
        todas_opcoes = sorted(set(
            campo.strip()
            for val in conc_alt["campos_alterados"]
            for campo in val.split("|")
            if campo.strip() and campo.strip() != "—"
        ))

        col_f1, col_f2, col_f3 = st.columns([2, 2, 1])

        with col_f1:
            campos_sel = st.multiselect(
                "Filtrar por campo alterado",
                options=todas_opcoes,
                default=[],
                placeholder="Todos os campos...",
            )

        with col_f2:
            status_opcoes = sorted(conc_alt["status_pagamento_atu"].unique().tolist())
            status_sel = st.multiselect(
                "Filtrar por status atual",
                options=status_opcoes,
                default=[],
                placeholder="Todos os status...",
            )

        with col_f3:
            st.markdown("<br>", unsafe_allow_html=True)
            var_neg = st.checkbox("Só variação negativa em CAR", value=False)

        # Aplicar filtros
        df_filtrado = conc_alt.copy()

        if campos_sel:
            df_filtrado = df_filtrado[
                df_filtrado["campos_alterados"].apply(
                    lambda x: any(c in x for c in campos_sel)
                )
            ]

        if status_sel:
            df_filtrado = df_filtrado[
                df_filtrado["status_pagamento_atu"].isin(status_sel)
            ]

        if var_neg:
            df_filtrado = df_filtrado[df_filtrado["var_contas_a_receber"] < 0]

        # Contador
        st.markdown(
            f"<p style='font-size:13px;color:#6B7A99;margin:8px 0 12px;'>"
            f"Exibindo <b>{len(df_filtrado):,}</b> de <b>{len(conc_alt):,}</b> cobranças alteradas</p>",
            unsafe_allow_html=True,
        )

        # Tabela
        COLUNAS_EXIBIR = [
            "id_cobranca",
            "campos_alterados",
            "status_pagamento_ant",
            "status_pagamento_atu",
            "valor_pago_ant",
            "valor_pago_atu",
            "var_valor_pago",
            "valor_excedente_pago_ant",
            "valor_excedente_pago_atu",
            "var_excedente_pago",
            "contas_a_receber_ant",
            "contas_a_receber_atu",
            "var_contas_a_receber",
        ]

        RENOMEAR = {
            "id_cobranca":              "ID Cobrança",
            "campos_alterados":         "Campos Alterados",
            "status_pagamento_ant":     f"Status ({label_ant})",
            "status_pagamento_atu":     f"Status ({label_atu})",
            "valor_pago_ant":           "Valor Pago Ant.",
            "valor_pago_atu":           "Valor Pago Atu.",
            "var_valor_pago":           "Var. Valor Pago",
            "valor_excedente_pago_ant": "Excedente Ant.",
            "valor_excedente_pago_atu": "Excedente Atu.",
            "var_excedente_pago":       "Var. Excedente",
            "contas_a_receber_ant":     "CAR Ant.",
            "contas_a_receber_atu":     "CAR Atu.",
            "var_contas_a_receber":     "Var. CAR",
        }

        df_exibir = df_filtrado[COLUNAS_EXIBIR].rename(columns=RENOMEAR)

        # Formatar colunas numéricas
        num_cols = [c for c in df_exibir.columns if c not in ("ID Cobrança","Campos Alterados",f"Status ({label_ant})",f"Status ({label_atu})")]
        df_fmt = df_exibir.copy()
        for c in num_cols:
            df_fmt[c] = df_fmt[c].apply(lambda v: f"R$ {v:,.2f}" if pd.notna(v) else "")

        st.dataframe(df_fmt, use_container_width=True, hide_index=True, height=420)

# ── Rodapé ─────────────────────────────────────────────────────────────────
st.markdown("<br><br>", unsafe_allow_html=True)
st.markdown(
    "<p style='text-align:center;font-size:12px;color:#A0AEC0;'>"
    "Conciliação de Carteira — Direitos Creditórios &nbsp;|&nbsp; "
    "Chave: <code>id_cobranca</code> &nbsp;|&nbsp; "
    "Campos: valor_cobranca · valor_pago · valor_excedente_pago · status_pagamento · contas_a_receber"
    "</p>",
    unsafe_allow_html=True,
)
